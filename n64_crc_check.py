#!/usr/bin/env python3
"""
Recursively compare ROM CRC32 values against a known CRC list stored in an .xlsx file.

Example:
  python3 n64_crc_check.py \
    --spreadsheet "/path/to/N64 UK Complete Game List.xlsx" \
    --rom-dir "/media/user/MiSTer_Data/games/N64"

Requirements:
  pip install openpyxl
"""
from __future__ import annotations

import argparse
import csv
import os
import sys
import zlib
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


def normalise_crc(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().upper()
    if text.startswith("0X"):
        text = text[2:]
    # Keep hex only
    text = "".join(ch for ch in text if ch in "0123456789ABCDEF")
    return text.zfill(8) if text else ""


def file_crc32(path: Path, chunk_size: int = 1024 * 1024) -> str:
    crc = 0
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(chunk_size), b""):
            crc = zlib.crc32(chunk, crc)
    return f"{crc & 0xFFFFFFFF:08X}"


def load_crc_map(spreadsheet: Path, sheet_name: str | None = None) -> tuple[dict[str, list[dict]], list[str]]:
    wb = load_workbook(spreadsheet, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    headers = [str(cell).strip() if cell is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    header_map = {name.upper(): idx for idx, name in enumerate(headers)}

    required = ["TITLE", "CRC"]
    missing = [name for name in required if name not in header_map]
    if missing:
        raise ValueError(f"Spreadsheet is missing required column(s): {', '.join(missing)}")

    title_idx = header_map["TITLE"]
    crc_idx = header_map["CRC"]
    game_id_idx = header_map.get("GAME ID")
    region_idx = header_map.get("REGION")
    media_title_idx = header_map.get("MEDIA TITLE")

    crc_map: dict[str, list[dict]] = defaultdict(list)
    duplicates: list[str] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        title = row[title_idx] if title_idx < len(row) else None
        crc = normalise_crc(row[crc_idx] if crc_idx < len(row) else None)
        if not crc:
            continue
        entry = {
            "title": "" if title is None else str(title).strip(),
            "game_id": "" if game_id_idx is None or game_id_idx >= len(row) or row[game_id_idx] is None else str(row[game_id_idx]).strip(),
            "region": "" if region_idx is None or region_idx >= len(row) or row[region_idx] is None else str(row[region_idx]).strip(),
            "media_title": "" if media_title_idx is None or media_title_idx >= len(row) or row[media_title_idx] is None else str(row[media_title_idx]).strip(),
            "crc": crc,
        }
        if crc_map[crc]:
            duplicates.append(crc)
        crc_map[crc].append(entry)

    return crc_map, sorted(set(duplicates))


def iter_rom_files(root: Path, extensions: set[str] | None = None):
    if extensions is None:
        extensions = {".z64", ".n64", ".v64", ".rom", ".bin", ".zip", ".7z", ".ndd"}
    for path in root.rglob("*"):
        if path.is_file() and path.suffix.lower() in extensions:
            yield path


def main() -> int:
    parser = argparse.ArgumentParser(description="Recursively compare ROM CRC32 values against a spreadsheet CRC list.")
    parser.add_argument("--spreadsheet", required=True, type=Path, help="Path to the .xlsx CRC list")
    parser.add_argument("--rom-dir", required=True, type=Path, help="Root directory to scan recursively")
    parser.add_argument("--sheet", default=None, help="Worksheet name (defaults to first sheet)")
    parser.add_argument("--report", default="crc_report.csv", type=Path, help="Output CSV report path")
    parser.add_argument(
        "--extensions",
        default=".z64,.n64,.v64,.rom,.bin,.zip,.7z,.ndd",
        help="Comma-separated file extensions to scan",
    )
    args = parser.parse_args()

    if not args.spreadsheet.is_file():
        print(f"Spreadsheet not found: {args.spreadsheet}", file=sys.stderr)
        return 2
    if not args.rom_dir.is_dir():
        print(f"ROM directory not found: {args.rom_dir}", file=sys.stderr)
        return 2

    exts = {ext.strip().lower() if ext.strip().startswith(".") else f".{ext.strip().lower()}"
            for ext in args.extensions.split(",") if ext.strip()}

    try:
        crc_map, duplicate_crcs = load_crc_map(args.spreadsheet, args.sheet)
    except Exception as exc:
        print(f"Failed to read spreadsheet: {exc}", file=sys.stderr)
        return 2

    scanned = 0
    matches = 0
    unknown = 0
    seen_crc_matches: set[str] = set()

    with args.report.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["status", "file", "crc32", "matched_title", "game_id", "region", "media_title", "notes"])

        for rom_path in sorted(iter_rom_files(args.rom_dir, exts)):
            scanned += 1
            try:
                crc = file_crc32(rom_path)
            except Exception as exc:
                writer.writerow(["ERROR", str(rom_path), "", "", "", "", "", str(exc)])
                continue

            entries = crc_map.get(crc, [])
            if entries:
                matches += 1
                seen_crc_matches.add(crc)
                note = "duplicate CRC in spreadsheet" if len(entries) > 1 else ""
                for entry in entries:
                    writer.writerow([
                        "MATCH",
                        str(rom_path),
                        crc,
                        entry["title"],
                        entry["game_id"],
                        entry["region"],
                        entry["media_title"],
                        note,
                    ])
            else:
                unknown += 1
                writer.writerow(["UNKNOWN", str(rom_path), crc, "", "", "", "", "CRC not found in spreadsheet"])

        missing = []
        for crc, entries in crc_map.items():
            if crc not in seen_crc_matches:
                for entry in entries:
                    missing.append(entry)
                    writer.writerow([
                        "MISSING_ON_SD",
                        "",
                        crc,
                        entry["title"],
                        entry["game_id"],
                        entry["region"],
                        entry["media_title"],
                        "",
                    ])

    print(f"Scanned files      : {scanned}")
    print(f"Matched CRCs       : {matches}")
    print(f"Unknown CRCs       : {unknown}")
    print(f"Missing from SD    : {len(missing)}")
    print(f"Duplicate CRC rows : {len(duplicate_crcs)}")
    print(f"CSV report written : {args.report}")

    if duplicate_crcs:
        print("\nDuplicate CRC values found in spreadsheet:")
        for crc in duplicate_crcs:
            print(f"  {crc}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
