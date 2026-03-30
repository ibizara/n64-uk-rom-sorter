#!/usr/bin/env python3
from __future__ import annotations

import binascii
import csv
import hashlib
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


WORKBOOK_PATH = Path("N64 UK Complete Game List.xlsx")
SHEET_NAME = "Sheet1"

IN_DIR = Path("IN_N64")
OUT_DIR = Path("OUT")
REPORTS_DIR = Path("reports")

EUR_DIR = OUT_DIR / "Europe"
USA_DIR = OUT_DIR / "USA"
JPN_DIR = OUT_DIR / "Japan"
INVALID_DIR = OUT_DIR / "INVALID"

DELETE_DUPLICATES = True


@dataclass
class RomEntry:
    title: str
    game_id: str
    crc32: str
    region: str


def normalise_crc(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().upper()
    text = text.replace("0X", "").replace("&H", "").replace(" ", "")
    text = "".join(ch for ch in text if ch in "0123456789ABCDEF")
    return text.zfill(8) if text else ""


def normalise_region(value: object) -> str:
    """
    Read the region directly from the Excel sheet and map it to
    Europe / USA / Japan / INVALID.
    """
    if value is None:
        return "INVALID"

    text = str(value).strip().upper()

    if not text:
        return "INVALID"

    # Europe / PAL
    if text in {"EUROPE", "EUR", "PAL", "UK", "ENGLISH", "EU"}:
        return "Europe"

    # USA / NTSC-U
    if text in {"USA", "US", "NTSC-U", "NTSC", "U"}:
        return "USA"

    # Japan / NTSC-J
    if text in {"JAPAN", "JPN", "JP", "NTSC-J", "J"}:
        return "Japan"

    return "INVALID"


def detect_columns(header_row: List[str]) -> Dict[str, int]:
    headers = {str(v).strip().lower(): i for i, v in enumerate(header_row)}

    def find(*candidates: str) -> Optional[int]:
        for cand in candidates:
            for header, idx in headers.items():
                if header == cand or cand in header:
                    return idx
        return None

    title_idx = find("title", "game name", "name")
    game_id_idx = find("game id", "serial", "id")
    region_idx = find("region")
    crc_idx = find("crc", "crc32")

    if title_idx is None or game_id_idx is None or region_idx is None or crc_idx is None:
        raise ValueError(f"Could not find required columns. Found headers: {list(headers.keys())}")

    return {
        "title": title_idx,
        "game_id": game_id_idx,
        "region": region_idx,
        "crc32": crc_idx,
    }


def load_rom_database(workbook_path: Path, sheet_name: str) -> Tuple[Dict[str, RomEntry], List[RomEntry]]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Worksheet is empty")

    header = ["" if v is None else str(v) for v in rows[0]]
    col = detect_columns(header)

    crc_map: Dict[str, RomEntry] = {}
    all_entries: List[RomEntry] = []

    for row in rows[1:]:
        title = "" if row[col["title"]] is None else str(row[col["title"]]).strip()
        game_id = "" if row[col["game_id"]] is None else str(row[col["game_id"]]).strip().upper()
        region = normalise_region(row[col["region"]])
        crc32 = normalise_crc(row[col["crc32"]])

        if not title or not game_id:
            continue

        entry = RomEntry(
            title=title,
            game_id=game_id,
            crc32=crc32,
            region=region,
        )
        all_entries.append(entry)

        if crc32 and crc32 not in crc_map:
            crc_map[crc32] = entry

    return crc_map, all_entries


def crc32_upper(path: Path) -> str:
    crc = 0
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            crc = binascii.crc32(chunk, crc)
    return f"{crc & 0xFFFFFFFF:08X}"


def sha1_upper(path: Path) -> str:
    h = hashlib.sha1()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest().upper()


def ensure_dirs() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    EUR_DIR.mkdir(parents=True, exist_ok=True)
    USA_DIR.mkdir(parents=True, exist_ok=True)
    JPN_DIR.mkdir(parents=True, exist_ok=True)
    INVALID_DIR.mkdir(parents=True, exist_ok=True)


def region_dir(region: str) -> Path:
    if region == "Europe":
        return EUR_DIR
    if region == "USA":
        return USA_DIR
    if region == "Japan":
        return JPN_DIR
    return INVALID_DIR


def unique_destination(base_dir: Path, stem: str, suffix: str) -> Path:
    candidate = base_dir / f"{stem}{suffix}"
    if not candidate.exists():
        return candidate

    n = 2
    while True:
        candidate = base_dir / f"{stem} ({n}){suffix}"
        if not candidate.exists():
            return candidate
        n += 1


def main() -> None:
    if not IN_DIR.exists():
        raise FileNotFoundError(f"Input folder not found: {IN_DIR}")

    ensure_dirs()
    crc_map, all_entries = load_rom_database(WORKBOOK_PATH, SHEET_NAME)

    seen_sha1: Dict[str, Path] = {}
    matched_crcs: set[str] = set()

    processed = 0
    matched = 0
    duplicates = 0
    invalid = 0

    sort_log_rows = []

    for path in sorted(IN_DIR.iterdir()):
        if not path.is_file():
            continue
        if path.suffix.lower() != ".n64":
            continue

        processed += 1

        try:
            crc = crc32_upper(path)
            sha1 = sha1_upper(path)

            entry = crc_map.get(crc)

            if entry is None:
                invalid += 1
                dest = unique_destination(INVALID_DIR, path.stem, path.suffix.lower())
                shutil.move(str(path), str(dest))
                sort_log_rows.append([path.name, crc, "INVALID", "", "", dest.as_posix()])
                continue

            if sha1 in seen_sha1:
                duplicates += 1
                if DELETE_DUPLICATES:
                    path.unlink()
                    status = "DUPLICATE_DELETED"
                    dest_text = "DELETED"
                else:
                    dest = unique_destination(region_dir(entry.region), entry.title, ".n64")
                    shutil.move(str(path), str(dest))
                    status = "DUPLICATE_KEPT"
                    dest_text = dest.as_posix()

                sort_log_rows.append([path.name, crc, status, entry.title, entry.game_id, dest_text])
                matched_crcs.add(crc)
                continue

            out_folder = region_dir(entry.region)
            dest = unique_destination(out_folder, entry.title, ".n64")

            shutil.move(str(path), str(dest))

            seen_sha1[sha1] = dest
            matched_crcs.add(crc)
            matched += 1

            sort_log_rows.append([path.name, crc, "MATCHED", entry.title, entry.game_id, dest.as_posix()])

        except Exception as exc:
            invalid += 1
            dest = unique_destination(INVALID_DIR, path.stem, path.suffix.lower())
            if path.exists():
                shutil.move(str(path), str(dest))
            sort_log_rows.append([path.name, "", f"ERROR: {exc}", "", "", dest.as_posix()])

    missing_rows = []
    for entry in all_entries:
        if entry.crc32 and entry.crc32 not in matched_crcs:
            missing_rows.append([entry.title, entry.game_id, entry.crc32, entry.region])

    with (REPORTS_DIR / "sort_log.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["source_file", "crc32", "status", "title", "game_id", "destination"])
        writer.writerows(sort_log_rows)

    with (REPORTS_DIR / "missing_from_collection.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["title", "game_id", "crc32", "region"])
        writer.writerows(missing_rows)

    print(f"Processed : {processed}")
    print(f"Matched   : {matched}")
    print(f"Duplicates: {duplicates}")
    print(f"Invalid   : {invalid}")
    print(f"Missing   : {len(missing_rows)}")
    print("Created   : OUT/, reports/")
    print("Reports   : reports/sort_log.csv, reports/missing_from_collection.csv")


if __name__ == "__main__":
    main()
